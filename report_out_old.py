import csv
import math
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00



def formatter_date(input_date):
    """
    Функия преобразует дату в нужный формат

    Args:
        input_date (str): Исходная данные

    Returns:
        (str): Дата в нужном формате
    """
    # return datetime.strptime(input_date, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')
    return input_date[:4]
    # input_date = input_date.split('-')
    # return input_date[0]

def exit_with_print(line):
    """
    Функция выводит строку и завершает программу

    Args:
        line (str): входная строка
    """
    print(line)
    exit()


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class Salary:
    """
    Класс, который хранит поля, связанные с зарплатой

    Attributes:
        salary_from (int or float): Минимальная граница оклада
        salary_to (int or float): Максимальная граница оклада
        salary_currency (str): Идентификатор валюты
        salary_ru (int or float): Валюта в рублях
    """

    def __init__(self, salary_from, salary_to, salary_currency):
        """
        В конструкторе устанавливаются основные поля зарплаты,
        а так же поле, конвертированной в рубль иностранной валюты

        Args:
            salary_from (str or float or int): Минимальная граница оклада
            salary_to (str or float or int): Максимальная граница оклада
            salary_currency (str or float or int): Идентификатор валюты

        >>> type(Salary(10000, 50000, 'RUR')).__name__
        'Salary'
        >>> Salary(10000, 50000, 'RUR').salary_from
        10000
        >>> Salary(10000, 50000, 'RUR').salary_to
        50000
        >>> Salary(10000, 50000, 'RUR').salary_currency
        'RUR'
        >>> Salary(10000, 50000, 'RUR').salary_ru
        30000
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        salary = Salary.currency_translate(salary_from, salary_to, salary_currency)
        self.salary_ru = int((salary[0] + salary[1]) / 2)

    @staticmethod
    def currency_translate(salary_from, salary_to, salary_currency):
        """
        Метод переводит зарплату в иностранной валюты в рубли

        Args:
            salary_from (str or float or int): Минимальная граница оклада
            salary_to (str or float or int): Максимальная граница оклада
            salary_currency (str or float or int): Идентификатор валюты

        Returns:
            (float, float): Кортеж, в котором хранится минимальная и максимальная зарплата в рублях

        >>> Salary.currency_translate('10000', '50000', 'RUR')
        (10000, 50000)
        >>> Salary.currency_translate('10000', '50000', 'AZN')
        (356800.0, 1784000.0)
        >>> Salary.currency_translate('8000', '20000', 'BYR')
        (191280.0, 478200.0)
        >>> Salary.currency_translate('500', '3000', 'EUR')
        (29950.0, 179700.0)
        >>> Salary.currency_translate('1500', '6000', 'GEL')
        (32609.999999999996, 130439.99999999999)
        >>> Salary.currency_translate('50000', '150000', 'KGS')
        (38000.0, 114000.0)
        >>> Salary.currency_translate(100000, 300000, 'KZT')
        (13000.0, 39000.0)
        >>> Salary.currency_translate(15000.50, 30000, 'UAH')
        (24600.0, 49200.0)
        >>> Salary.currency_translate(2000, 4000, 'USD')
        (121320.0, 242640.0)
        >>> Salary.currency_translate(2000000, 4000000, 'UZS')
        (11000.0, 22000.0)
        """
        salary_from = int(math.trunc(float(salary_from))) * currency_to_rub[salary_currency]
        salary_to = int(math.trunc(float(salary_to))) * currency_to_rub[salary_currency]
        return salary_from, salary_to


class Vacancy:
    """
    Класс, который хранит поля, связанные с вакансией

    Attributes:
        name (str): Название вакансий
        salary (Salary): Объект Salary
        area_name (str): Название региона
        published_at (str): Дата публикации
    """
    def __init__(self, dictionary):
        """
        В конструкторе устанавливаются основные поля для вакансии

        Args:
            dictionary (dict): Словарь, содержащий данные о вакансии

        >>> type(Vacancy({'name': 'IT аналитик', 'salary_from': '35000.0', 'salary_to': '45000.0', 'salary_currency': 'RUR', \
        'area_name': 'Санкт-Петербург', 'published_at': '2007-12-03T17:34:36+0300'})).__name__
        'Vacancy'
        >>> Vacancy({'name': 'IT аналитик', 'salary_from': '35000.0', 'salary_to': '45000.0', 'salary_currency': 'RUR', \
        'area_name': 'Санкт-Петербург', 'published_at': '2007-12-03T17:34:36+0300'}).name
        'IT аналитик'
        >>> type(Vacancy({'name': 'IT аналитик', 'salary_from': '35000.0', 'salary_to': '45000.0', 'salary_currency': 'RUR', \
        'area_name': 'Санкт-Петербург' , 'published_at': '2007-12-03T17:34:36+0300'}).salary).__name__
        'Salary'
        >>> Vacancy({'name': 'IT аналитик', 'salary_from': '35000.0', 'salary_to': '45000.0', 'salary_currency': 'RUR', \
        'area_name': 'Санкт-Петербург', 'published_at': '2007-12-03T17:34:36+0300'}).area_name
        'Санкт-Петербург'
        >>> Vacancy({'name': 'IT аналитик', 'salary_from': '35000.0', 'salary_to': '45000.0', 'salary_currency': 'RUR', \
        'area_name': 'Санкт-Петербург', 'published_at': '2007-12-03T17:34:36+0300'}).published_at
        '2007-12-03T17:34:36+0300'
        """
        self.name = dictionary['name']
        self.salary = Salary(dictionary['salary_from'], dictionary['salary_to'], dictionary['salary_currency'])
        self.area_name = dictionary['area_name']
        self.published_at = dictionary['published_at']


class DataSet:
    """
    Класс отвечает за чтение и обработку данных из CSV файла

    Attributes:
        file_name (str): Имя файла
        vacancies_objects (list): Список из объектов Vacancy
    """
    def __init__(self, file_name):
        """
        В конструкторе устанавливаются основные поля для набора данных

        Args:
            file_name (str): Имя входного файла
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.prepare_data(file_name)

    @staticmethod
    def read_csv(file_name):
        """
        Метод считывает данные из CSV файла

        Args:
            file_name (str): Имя входного файла

        Returns:
            (list, list): Кортеж из списка названий колонок и
            списка непосредственно данных о вакансиях
        """
        reader_csv = csv.reader(open(file_name, encoding='utf_8_sig'))
        list_data = [x for x in reader_csv]
        if len(list_data) == 0:
            exit_with_print("Пустой файл")
        if len(list_data) == 1:
            exit_with_print("Нет данных")
        columns = list_data[0]
        vacancies = [x for x in list_data[1:] if len(x) == len(columns) and x.count('') == 0]
        return columns, vacancies

    @staticmethod
    def prepare_data(file_name):
        """
        Метод обрабатывает данные из CSV файла и преобразует их в список вакансий

        Args:
            file_name (str): Имя входного файла

        Returns:
            (list): Список с объектами Vacancy
        """
        columns, vacancies = DataSet.read_csv(file_name)
        list_vacancies = []
        for row in vacancies:
            vacancy_dict = {}
            for i in range(len(row)):
                vacancy_dict[columns[i]] = row[i]
            list_vacancies.append(Vacancy(vacancy_dict))
        return list_vacancies


class InputConnect:
    """
    Класс отвечает за работу с входными параметрами и за печать набора данных
    """
    def __init__(self):
        """
        Конструктор запускает метод, получающий входные данные, создает
        набор данных и запускает метод по печати этого набора
        """
        params = InputConnect.get_params()
        data_set = DataSet(params[0])
        InputConnect.print_data(data_set.vacancies_objects, params[1])

    @staticmethod
    def get_params():
        """
        Метод получает входные данные

        Returns:
            (str, str): Кортеж входных параметров
        """
        file_name = input('Введите название файла: ')
        job_name = input('Введите название профессии: ')
        return file_name, job_name

    @staticmethod
    def print_data(list_vacancies, job_name):
        """
        Метод обрабатывает набор данных и печатает их. Так же
        метод запускает формирование графиков и отчетов

        Args:
            list_vacancies (list): Список с данными о вакансиях
            job_name (str): Вакансия, по которой будет вестись статистика
        """
        years = set()
        for vacancy in list_vacancies:
            years.add(int(formatter_date(vacancy.published_at)))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))

        salary_by_years = {year: [] for year in years}
        vacs_by_years = {year: 0 for year in years}
        job_salary_by_years = {year: [] for year in years}
        job_count_by_years = {year: 0 for year in years}
        area_dict = {}
        vacs_dict = {}

        for vacancy in list_vacancies:
            year = int(formatter_date(vacancy.published_at))
            salary_by_years[year].append(vacancy.salary.salary_ru)
            vacs_by_years[year] += 1
            if job_name in vacancy.name:
                job_salary_by_years[year].append(vacancy.salary.salary_ru)
                job_count_by_years[year] += 1
            if vacancy.area_name in area_dict:
                area_dict[vacancy.area_name].append(vacancy.salary.salary_ru)
            else:
                area_dict[vacancy.area_name] = [vacancy.salary.salary_ru]
            if vacancy.area_name in vacs_dict:
                vacs_dict[vacancy.area_name] += 1
            else:
                vacs_dict[vacancy.area_name] = 1

        salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                           salary_by_years.items()}
        job_salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                               job_salary_by_years.items()}

        area_list = area_dict.items()
        area_list = [x for x in area_list if len(x[1]) / len(list_vacancies) > 0.01]
        area_list = sorted(area_list, key=lambda x: sum(x[1]) / len(x[1]), reverse=True)
        salary_by_cities = {x[0]: int(sum(x[1]) / len(x[1])) for x in area_list[0: min(len(area_list), 10)]}

        vacs_count = {x: round(y / len(list_vacancies), 4) for x, y in vacs_dict.items()}
        vacs_count = {key: value for key, value in vacs_count.items() if value >= 0.01}
        vacs_by_cities = dict(sorted(vacs_count.items(), key=lambda x: x[1], reverse=True))
        vacs_by_cities = dict(list(vacs_by_cities.items())[:10])

        print('Динамика уровня зарплат по годам:', salary_by_years)
        print('Динамика количества вакансий по годам:', vacs_by_years)
        print('Динамика уровня зарплат по годам для выбранной профессии:', job_salary_by_years)
        print('Динамика количества вакансий по годам для выбранной профессии:', job_count_by_years)
        print('Уровень зарплат по городам (в порядке убывания):', salary_by_cities)
        print('Доля вакансий по городам (в порядке убывания):', vacs_by_cities)

        exit()
        data_list = [salary_by_years, vacs_by_years, job_salary_by_years, job_count_by_years, salary_by_cities,
                     vacs_by_cities]

        Report(data_list, job_name)


class Report:
    """
    Класс отвечает за формирование графика и отчетов в виде xlsx, pdf
    """
    # Поля необходимые для метода generate_pdf
    heads1 = []
    heads2 = []
    job_name = ''
    data_list = []

    def __init__(self, data_list, job_name):
        """
        В конструкторе устанавливаются поля, необходимые для
        метода generate_pdf, а так же происходит запуск остальных
        методов класса

        Args:
            data_list (list): Список списков значений статистики
            job_name (str): Вакансия, по которой будет вестись статистика
        """
        Report.job_name = job_name
        Report.data_list = data_list
        Report.generate_excel()
        Report.generate_image()
        Report.generate_pdf()

    @staticmethod
    def as_text(line):
        """
        Метод преобразует переменную в строку

        Args:
            line: Переменная, которую нужно перевести в строку

        Returns:
            (str): Переменная в виде строки
        """
        if line is None:
            return ''
        return str(line)

    @staticmethod
    def create_sheet2(sheet2, data_list):
        """
        Метод формирует вторую вкладу таблицы

        Args:
            sheet2: Объект вкладки WorkSheet
            data_list (list): Список списков значений статистики
        """
        heads2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
        Report.heads2 = heads2

        for i in range(len(heads2)):
            sheet2.cell(row=1, column=(i + 1), value=heads2[i]).font = Font(bold=True)

        salary_key = list(data_list[4])
        vacs_key = list(data_list[5])
        for i in range(len(salary_key)):
            sheet2.append([salary_key[i], data_list[4][salary_key[i]], vacs_key[i], data_list[5][vacs_key[i]]])
            sheet2['D' + str(i + 2)].number_format = FORMAT_PERCENTAGE_00

        sheet2.insert_cols(3, 1)

    @staticmethod
    def generate_excel():
        """
        Метод, который формирует отчет в виде xlsx

        """
        data_list = Report.data_list
        job_name = Report.job_name
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = 'Статистика по годам'
        heads1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {job_name}', 'Количество вакансий',
                  f'Количество вакансий - {job_name}']
        Report.heads1 = heads1

        for i in range(len(heads1)):
            sheet1.cell(row=1, column=(i + 1), value=heads1[i]).font = Font(bold=True)

        for key, value in data_list[0].items():
            sheet1.append([key, value, data_list[2][key], data_list[1][key], data_list[3][key]])

        thin = Side(border_style='thin', color='000000')
        for column in sheet1.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet1.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        sheet2 = wb.create_sheet('Статистика по городам')
        Report.create_sheet2(sheet2, data_list)

        for column in sheet2.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        wb.save('report.xlsx')

    @staticmethod
    def generate_image():
        """
        Метод создает графики по статистике

        """
        data_list = Report.data_list
        job_name = Report.job_name
        fig = plt.figure()
        width = 0.4
        x_nums = np.arange(len(data_list[0].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        ax = fig.add_subplot(221)
        ax.set_title('Уровень зарплат по годам')
        ax.bar(x_list1, data_list[0].values(), width, label='средняя з/п')
        ax.bar(x_list2, data_list[2].values(), width, label=f'з/п {job_name}')
        ax.set_xticks(x_nums, data_list[0].keys(), rotation='vertical')
        ax.legend(fontsize=8)
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(True, axis='y')

        x_nums = np.arange(len(data_list[1].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2
        ax = fig.add_subplot(222)
        ax.set_title('Количество вакансий по годам')
        ax.bar(x_list1, data_list[1].values(), width, label='Количество вакансий')
        ax.bar(x_list2, data_list[3].values(), width, label=f'Количество вакансий \n{job_name}')
        ax.set_xticks(x_nums, data_list[1].keys(), rotation='vertical')
        ax.legend(fontsize=8)
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(True, axis='y')

        y_nums = np.arange((len(data_list[4].keys())))
        ax = fig.add_subplot(223)
        ax.set_title('Уровень зарплат по городам')
        data = [line.replace(' ', '\n').replace('-', '-\n') for line in list(data_list[4].keys())]
        ax.barh(y_nums, data_list[4].values(), align='center')
        ax.set_yticks(y_nums, data)
        ax.tick_params(axis='y', labelsize=6)
        ax.tick_params(axis='x', labelsize=8)

        ax = fig.add_subplot(224)
        ax.set_title('Доля вакансий по городам')
        values = list(data_list[5].values())
        name = list(data_list[5].keys())
        values.insert(0, 1 - sum(values))
        name.insert(0, 'Другие')
        ax.pie(values, labels=name, textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png', dpi=200)

    @staticmethod
    def generate_pdf():
        """
        Метод формирует отчет в виде pdf

        """
        salary_key = list(Report.data_list[4])
        vacs_key = list(Report.data_list[5])
        vacs_by_cities = [str('{0:.2%}'.format(float(x))).replace('.', ',') for x in list(Report.data_list[5].values())]
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'heads1': Report.heads1, 'job': Report.job_name,
                                        'salary_by_years': Report.data_list[0],
                                        'job_salary_by_years': Report.data_list[2],
                                        'vacs_by_years': Report.data_list[1],
                                        'job_count_by_years': Report.data_list[3],
                                        'heads2': Report.heads2, 'salary_key': salary_key, 'vacs_key': vacs_key,
                                        'salary_by_cities': Report.data_list[4],
                                        'vacs_by_cities': vacs_by_cities})

        # with open("my_new_file.html", "w", encoding='utf-8') as fh:
        #     fh.write(pdf_template)
        # exit()

        config = pdfkit.configuration(wkhtmltopdf=r'C:\source\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={'enable-local-file-access': None})


