import csv
import math
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader


def exit_with_print(line):
    print(line)
    exit()


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        salary = Salary.currency_translate(salary_from, salary_to, salary_currency)
        self.salary_ru = int((salary[0] + salary[1]) / 2)

    @staticmethod
    def currency_translate(salary_from, salary_to, salary_currency):
        salary_from = int(math.trunc(float(salary_from))) * currency_to_rub[salary_currency]
        salary_to = int(math.trunc(float(salary_to))) * currency_to_rub[salary_currency]
        return salary_from, salary_to


class Vacancy:
    def __init__(self, dictionary):
        self.name = dictionary['name']
        self.salary = Salary(dictionary['salary_from'], dictionary['salary_to'], dictionary['salary_currency'])
        self.area_name = dictionary['area_name']
        self.published_at = dictionary['published_at']


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.prepare_data(file_name)

    @staticmethod
    def read_csv(file_name):
        reader_csv = csv.reader(open(file_name, encoding='utf_8_sig'))
        list_data = [x for x in reader_csv]
        if len(list_data) == 0:
            exit_with_print("Пустой файл")
        if len(list_data) == 1:
            exit_with_print("Нет данных")
        columns = list_data[0]
        vacancies = [x for x in list_data[1:] if len(x) == len(columns) and x.count('') == 0]
        return columns, vacancies

    @staticmethod
    def prepare_data(file_name):
        columns, vacancies = DataSet.read_csv(file_name)
        list_vacancies = []
        for row in vacancies:
            vacancy_dict = {}
            for i in range(len(row)):
                vacancy_dict[columns[i]] = row[i]
            list_vacancies.append(Vacancy(vacancy_dict))
        return list_vacancies


class InputConnect:
    def __init__(self):
        params = InputConnect.get_params()
        data_set = DataSet(params[0])
        InputConnect.print_data(data_set.vacancies_objects, params[1])

    @staticmethod
    def get_params():
        file_name = input('Введите название файла: ')
        job_name = input('Введите название профессии: ')
        return file_name, job_name

    @staticmethod
    def print_data(list_vacancies, job_name):
        years = set()
        for vacancy in list_vacancies:
            years.add(int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))

        salary_by_years = {year: [] for year in years}
        vacs_by_years = {year: 0 for year in years}
        job_salary_by_years = {year: [] for year in years}
        job_count_by_years = {year: 0 for year in years}
        area_dict = {}
        vacs_dict = {}

        for vacancy in list_vacancies:
            year = int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
            salary_by_years[year].append(vacancy.salary.salary_ru)
            vacs_by_years[year] += 1
            if job_name in vacancy.name:
                job_salary_by_years[year].append(vacancy.salary.salary_ru)
                job_count_by_years[year] += 1
            if vacancy.area_name in area_dict:
                area_dict[vacancy.area_name].append(vacancy.salary.salary_ru)
            else:
                area_dict[vacancy.area_name] = [vacancy.salary.salary_ru]
            if vacancy.area_name in vacs_dict:
                vacs_dict[vacancy.area_name] += 1
            else:
                vacs_dict[vacancy.area_name] = 1

        salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                           salary_by_years.items()}
        job_salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                               job_salary_by_years.items()}

        area_list = area_dict.items()
        area_list = [x for x in area_list if len(x[1]) / len(list_vacancies) > 0.01]
        area_list = sorted(area_list, key=lambda x: sum(x[1]) / len(x[1]), reverse=True)
        salary_by_cities = {x[0]: int(sum(x[1]) / len(x[1])) for x in area_list[0: min(len(area_list), 10)]}

        vacs_count = {x: round(y / len(list_vacancies), 4) for x, y in vacs_dict.items()}
        vacs_count = {key: value for key, value in vacs_count.items() if value >= 0.01}
        vacs_by_cities = dict(sorted(vacs_count.items(), key=lambda x: x[1], reverse=True))
        vacs_by_cities = dict(list(vacs_by_cities.items())[:10])

        data_list = [salary_by_years, vacs_by_years, job_salary_by_years, job_count_by_years, salary_by_cities,
                     vacs_by_cities]

        print('Динамика уровня зарплат по годам:', salary_by_years)
        print('Динамика количества вакансий по годам:', vacs_by_years)
        print('Динамика уровня зарплат по годам для выбранной профессии:', job_salary_by_years)
        print('Динамика количества вакансий по годам для выбранной профессии:', job_count_by_years)
        print('Уровень зарплат по городам (в порядке убывания):', salary_by_cities)
        print('Доля вакансий по городам (в порядке убывания):', vacs_by_cities)

        Report(data_list, job_name)


class Report:
    # Поля необходимые для метода generate_pdf
    heads1 = []
    heads2 = []
    job_name = ''
    data_list = []

    def __init__(self, data_list, job_name):
        Report.job_name = job_name
        Report.data_list = data_list
        Report.generate_excel(data_list, job_name)
        Report.generate_image(data_list, job_name)
        Report.generate_pdf()

    @staticmethod
    def as_text(line):
        if line is None:
            return ''
        return str(line)

    @staticmethod
    def create_sheet2(sheet2, data_list):
        heads2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
        Report.heads2 = heads2

        for i in range(len(heads2)):
            sheet2.cell(row=1, column=(i + 1), value=heads2[i]).font = Font(bold=True)

        salary_key = list(data_list[4])
        vacs_key = list(data_list[5])
        for i in range(len(salary_key)):
            sheet2.append([salary_key[i], data_list[4][salary_key[i]], vacs_key[i], data_list[5][vacs_key[i]]])
            sheet2['D' + str(i + 2)].number_format = FORMAT_PERCENTAGE_00

        sheet2.insert_cols(3, 1)

    @staticmethod
    def generate_excel(data_list, job_name):
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = 'Статистика по годам'
        heads1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {job_name}', 'Количество вакансий',
                  f'Количество вакансий - {job_name}']
        Report.heads1 = heads1

        for i in range(len(heads1)):
            sheet1.cell(row=1, column=(i + 1), value=heads1[i]).font = Font(bold=True)

        for key, value in data_list[0].items():
            sheet1.append([key, value, data_list[2][key], data_list[1][key], data_list[3][key]])

        thin = Side(border_style='thin', color='000000')
        for column in sheet1.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet1.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        sheet2 = wb.create_sheet('Статистика по городам')
        Report.create_sheet2(sheet2, data_list)

        for column in sheet2.columns:
            length = max(len(Report.as_text(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)

        wb.save('report.xlsx')

    @staticmethod
    def generate_image(data_list, job_name):
        fig = plt.figure()
        width = 0.4
        x_nums = np.arange(len(data_list[0].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2

        ax = fig.add_subplot(221)
        ax.set_title('Уровень зарплат по годам')
        ax.bar(x_list1, data_list[0].values(), width, label='средняя з/п')
        ax.bar(x_list2, data_list[2].values(), width, label=f'з/п {job_name}')
        ax.set_xticks(x_nums, data_list[0].keys(), rotation='vertical')
        ax.legend(fontsize=8)
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(True, axis='y')

        x_nums = np.arange(len(data_list[1].keys()))
        x_list1 = x_nums - width / 2
        x_list2 = x_nums + width / 2
        ax = fig.add_subplot(222)
        ax.set_title('Количество вакансий по годам')
        ax.bar(x_list1, data_list[1].values(), width, label='Количество вакансий')
        ax.bar(x_list2, data_list[3].values(), width, label=f'Количество вакансий \n{job_name}')
        ax.set_xticks(x_nums, data_list[1].keys(), rotation='vertical')
        ax.legend(fontsize=8)
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(True, axis='y')

        y_nums = np.arange((len(data_list[4].keys())))
        ax = fig.add_subplot(223)
        ax.set_title('Уровень зарплат по городам')
        data = [line.replace(' ', '\n').replace('-', '-\n') for line in list(data_list[4].keys())]
        ax.barh(y_nums, data_list[4].values(), align='center')
        ax.set_yticks(y_nums, data)
        ax.tick_params(axis='y', labelsize=6)
        ax.tick_params(axis='x', labelsize=8)

        ax = fig.add_subplot(224)
        ax.set_title('Доля вакансий по городам')
        values = list(data_list[5].values())
        name = list(data_list[5].keys())
        values.insert(0, 1 - sum(values))
        name.insert(0, 'Другие')
        ax.pie(values, labels=name, textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png', dpi=200)

    @staticmethod
    def generate_pdf():
        salary_key = list(Report.data_list[4])
        vacs_key = list(Report.data_list[5])
        vacs_by_cities = [str('{0:.2%}'.format(float(x))).replace('.', ',') for x in list(Report.data_list[5].values())]
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'heads1': Report.heads1, 'job': Report.job_name,
                                        'salary_by_years': Report.data_list[0],
                                        'job_salary_by_years': Report.data_list[2],
                                        'vacs_by_years': Report.data_list[1],
                                        'job_count_by_years': Report.data_list[3],
                                        'heads2': Report.heads2, 'salary_key': salary_key, 'vacs_key': vacs_key,
                                        'salary_by_cities': Report.data_list[4],
                                        'vacs_by_cities': vacs_by_cities})

        # with open("my_new_file.html", "w", encoding='utf-8') as fh:
        #     fh.write(pdf_template)
        # exit()

        config = pdfkit.configuration(wkhtmltopdf=r'C:\source\wkhtmltox\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={'enable-local-file-access': None})
